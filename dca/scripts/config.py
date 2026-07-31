"""
00_config.py — Configuración central del pipeline Tomillo × Fusarium.

Define rutas, constantes, mapeo de concentraciones y funciones compartidas
para el parsing de las hojas del archivo Excel.
"""

from pathlib import Path
import pandas as pd
import numpy as np
from scipy import stats

# ─── Rutas ────────────────────────────────────────────────────────────
# RAIZ = raíz del repositorio; cada diseño (dca/, bdca/, factorial/) vive en
# su propia carpeta con datos crudos en datos_crudos/<diseño>/.
RAIZ = Path(__file__).resolve().parent.parent.parent
EXCEL_ORIGINAL = RAIZ / "datos_crudos" / "dca" / "datos-proyectos tomillo-fusarium.xlsx"
DIR_DATOS = RAIZ / "dca" / "datos_procesados"
DIR_RESULTADOS = RAIZ / "dca" / "resultados"
DIR_FIGURAS = DIR_RESULTADOS / "figuras"
DIR_TABLAS = DIR_RESULTADOS / "tablas"
DIR_REPORTES = DIR_RESULTADOS / "reportes"
DIR_EXCEL = DIR_RESULTADOS / "excel"  # tablas Excel consolidadas

# asegurar que existan al importar
for d in [DIR_DATOS, DIR_FIGURAS, DIR_TABLAS, DIR_REPORTES, DIR_EXCEL]:
    d.mkdir(parents=True, exist_ok=True)

# ─── Constantes ──────────────────────────────────────────────────────
SEMILLA_ALEATORIA = 42

# ─── Constantes de figuras (publicación científica) ──────────────────
FIGURA_DPI = 300
FIGURA_DPI_BAJO = 200  # para figuras muy grandes (dendrogramas, etc.)

# Paleta de colores para métodos de extracción (accesible, daltónico-safe)
COLOR_MET = {
    "maceracion": "#2e86ab",   # azul petróleo
    "soxhlet": "#a23b72",      # púrpura
    "ultrasonido": "#f18f01",  # naranja
}

# Paleta para gradiente de concentración (mapa de grises → secuencial)
# Se usan 5 colores para hasta 5 niveles: control → 0.2 → 1.0 → 5.0
COLOR_CONC = ["#b3b3b3", "#7ba0b4", "#4a7c9b", "#1a4d6b"]

# Etiquetas de visualización para métodos
LABEL_MET = {
    "maceracion": "Maceración",
    "soxhlet": "Soxhlet",
    "ultrasonido": "Ultrasonido",
}

# Etiquetas de concentración para ejes
LABEL_CONC_ORIG = {5.0: "5.0", 1.0: "1.0", 0.2: "0.2", 0.0: "Control"}

# Variables de respuesta
VAR_INH = "porcentaje_inhibicion"        # %INH crecimiento (crudo)
VAR_INH_LOG10 = "porcentaje_inhibicion_log10"  # %INH conidias (escala hoja)
VAR_CONIDIAS = "conidias_log10"          # log10(conidias/mL)
VAR_RENDIMIENTO = "rendimiento_pct"       # rendimiento (%)


def setup_figure_style(dpi=FIGURA_DPI, font_scale=1.0):
    """Configura matplotlib rcParams para figuras de publicación científica.

    Parámetros
    ----------
    dpi : int
        DPI base para figuras (default: 300).
    font_scale : float
        Factor de escala para tamaños de fuente (default: 1.0).

    Uso:
        from config import setup_figure_style
        setup_figure_style()
    """
    import matplotlib as mpl
    mpl.rcParams.update({
        "figure.dpi": dpi,
        "savefig.dpi": dpi,
        "font.size": 12 * font_scale,
        "axes.titlesize": 13 * font_scale,
        "axes.labelsize": 12 * font_scale,
        "xtick.labelsize": 10 * font_scale,
        "ytick.labelsize": 10 * font_scale,
        "legend.fontsize": 10 * font_scale,
        "legend.title_fontsize": 11 * font_scale,
        "figure.titlesize": 14 * font_scale,
        "lines.linewidth": 1.5,
        "lines.markersize": 6,
        "figure.facecolor": "white",
        "axes.facecolor": "white",
        "axes.edgecolor": ".3",
        "axes.grid": False,
        "grid.alpha": 0.3,
        "grid.linestyle": "--",
        "xtick.major.width": 0.8,
        "ytick.major.width": 0.8,
        "axes.spines.top": False,
        "axes.spines.right": False,
        "figure.constrained_layout.use": False,  # usamos tight_layout explícito
    })


def save_figure_pub(fig, filename, dir_fig=None, dpi=None, clean=True):
    """Guarda una figura en versión normal + clean (sin títulos, para publicación).

    La versión clean omite títulos de ejes y del figure, útil cuando el título
    va en el caption del artículo.

    Parámetros
    ----------
    fig : matplotlib.figure.Figure
        Figura a guardar.
    filename : str
        Nombre del archivo (ej. 'obj1_rendimiento.png').
    dir_fig : Path, optional
        Directorio de salida. Default: DIR_FIGURAS.
    dpi : int, optional
        DPI de exportación. Default: FIGURA_DPI.
    clean : bool
        Si True, genera también versión _clean.png sin títulos.
    """
    if dir_fig is None:
        dir_fig = DIR_FIGURAS
    if dpi is None:
        dpi = FIGURA_DPI

    # Versión normal
    fig.tight_layout()
    fig.savefig(dir_fig / filename, dpi=dpi, bbox_inches="tight")

    # Versión clean (sin títulos)
    if clean:
        # Preservar títulos originales para restaurar después
        original_titles = {}
        for i, ax in enumerate(fig.axes):
            original_titles[i] = ax.get_title()
            ax.set_title("")

        # También figure.suptitle si existe
        orig_suptitle = fig._suptitle  # guardar referencia
        fig.suptitle("")

        clean_name = filename.replace(".png", "_clean.png")
        fig.tight_layout()
        fig.savefig(dir_fig / clean_name, dpi=dpi, bbox_inches="tight")

        # Restaurar títulos (por si la figura se usa después)
        for i, ax in enumerate(fig.axes):
            if i in original_titles and original_titles[i]:
                ax.set_title(original_titles[i])
        if orig_suptitle:
            fig.suptitle(orig_suptitle.get_text() if orig_suptitle else "")

    plt.close(fig)


def styled_excel_export(filepath, sheet_data, sheet_order=None):
    """Escribe un libro Excel con formato de publicación científica.

    Parámetros
    ----------
    filepath : Path
        Ruta del archivo .xlsx a crear/sobrescribir.
    sheet_data : dict[str, pd.DataFrame]
        Diccionario {nombre_hoja: DataFrame} con los datos a exportar.
    sheet_order : list[str], optional
        Orden de las hojas. Si es None, usa el orden del dict.

    Retorna
    -------
    Path
        Ruta del archivo creado.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Eliminar hoja por defecto
    wb.remove(wb.active)

    sheets = sheet_order if sheet_order else list(sheet_data.keys())

    # Estilos
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )
    band_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    for sheet_name in sheets:
        df = sheet_data.get(sheet_name)
        if df is None or len(df) == 0:
            continue

        ws = wb.create_sheet(title=sheet_name[:31])  # Excel max 31 chars

        # Escribir encabezados
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Escribir datos
        for row_idx, (_, row) in enumerate(df.iterrows()):
            for col_idx, col_name in enumerate(df.columns, 1):
                val = row[col_name]
                cell = ws.cell(row=row_idx + 2, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")

                # Bandas alternas
                if row_idx % 2 == 1:
                    cell.fill = band_fill

                # Formato numérico
                if isinstance(val, float):
                    # p-valores: 4 decimales
                    if col_name.lower().startswith("p") or "p_val" in col_name.lower():
                        if val < 0.0001:
                            cell.value = "<0.0001"
                            cell.number_format = "@"
                        else:
                            cell.number_format = "0.0000"
                    else:
                        cell.number_format = "0.00"

        # Autoajustar ancho de columnas
        for col_idx, col_name in enumerate(df.columns, 1):
            max_len = len(str(col_name))
            for row_idx in range(min(len(df), 50)):  # muestra primeras 50 filas
                val = str(df.iloc[row_idx, col_idx - 1]) if df.iloc[row_idx, col_idx - 1] is not None else ""
                max_len = max(max_len, len(val))
            # Limitar ancho y dejar espacio
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)

        # Fijar primera fila
        ws.freeze_panes = "A2"

    wb.save(filepath)
    return filepath


# ─── Funciones de diagnóstico compartidas ───────────────────────────


def diagnostic_kmo_bartlett(data, var_names=None):
    """Calcula KMO (Kaiser-Meyer-Olkin) y test de Bartlett de esfericidad.

    KMO mide si los datos son adecuados para PCA (> 0.6 aceptable, > 0.8 excelente).
    Bartlett prueba H0: matriz de correlación = identidad (rechazar H0 es deseable).

    Parámetros
    ----------
    data : pd.DataFrame
        Datos numéricos sin NaNs.
    var_names : list, optional
        Nombres de variables a incluir. Si None, usa todas.

    Retorna
    -------
    dict con 'kmo_total', 'kmo_por_variable', 'bartlett_chi2', 'bartlett_p', 'determinante'.
    """
    from sklearn.preprocessing import StandardScaler

    if var_names:
        X = data[var_names].dropna().values
    else:
        X = data.values

    # Estandarizar
    X_scaled = StandardScaler().fit_transform(X)

    # Matriz de correlación
    corr = np.corrcoef(X_scaled.T)

    # KMO: ratio de correlaciones al cuadrado sobre correlaciones parciales al cuadrado
    # Inversa de la matriz de correlación → precision matrix
    try:
        corr_inv = np.linalg.inv(corr)
    except np.linalg.LinAlgError:
        # Si es singular, intentar pseudoinversa
        corr_inv = np.linalg.pinv(corr)

    # Matriz de correlaciones parciales ≈ -corr_inv / sqrt(diag(corr_inv) outer diag(corr_inv))
    partial_corr = -corr_inv
    d = np.sqrt(np.diag(corr_inv))
    partial_corr = partial_corr / np.outer(d, d)
    np.fill_diagonal(partial_corr, 0)

    # KMO por variable
    partial_sq = partial_corr ** 2
    corr_sq = corr ** 2
    np.fill_diagonal(corr_sq, 0)

    kmo_per_var = np.sum(corr_sq, axis=1) / (np.sum(corr_sq, axis=1) + np.sum(partial_sq, axis=1))

    # KMO total
    kmo_total = np.sum(corr_sq) / (np.sum(corr_sq) + np.sum(partial_sq))

    # Determinante de la matriz de correlación
    try:
        det = np.linalg.det(corr)
    except np.linalg.LinAlgError:
        det = np.nan

    # Bartlett de esfericidad
    n = X_scaled.shape[0]
    p = X_scaled.shape[1]
    chi2 = -((n - 1) - (2 * p + 5) / 6) * np.log(det) if det > 0 else np.nan
    df_bart = p * (p - 1) / 2
    p_val = stats.chi2.sf(chi2, df_bart) if not np.isnan(chi2) else np.nan

    return {
        "kmo_total": kmo_total,
        "kmo_por_variable": dict(zip(
            (var_names if var_names else [f"V{i+1}" for i in range(p)]),
            kmo_per_var
        )),
        "bartlett_chi2": chi2,
        "bartlett_df": df_bart,
        "bartlett_p": p_val,
        "determinante": det,
    }


def diagnostic_durbin_watson(residuos):
    """Calcula el estadístico Durbin-Watson para independencia de residuos.

    DW ≈ 2 → sin autocorrelación. DW < 1 o > 3 → preocupante.
    """
    diff = np.diff(residuos)
    dw = np.sum(diff ** 2) / np.sum(residuos ** 2)
    return dw


def diagnostic_breusch_pagan(modelo, exog):
    """Test de Breusch-Pagan para heterocedasticidad.

    H0: homocedasticidad. p < 0.05 → evidencia de heterocedasticidad.
    """
    residuos = modelo.resid
    n = len(residuos)
    residuos_cuad = residuos ** 2

    # Regresión auxiliar: residuos² ~ exog
    from statsmodels.regression.linear_model import OLS
    aux_model = OLS(residuos_cuad, exog).fit()
    r2 = aux_model.rsquared
    lm_stat = n * r2
    k = exog.shape[1] - 1  # número de regresores (sin constante)
    p_val = stats.chi2.sf(lm_stat, k)

    return {"lm_stat": lm_stat, "p_val": p_val, "df": k}


def diagnostic_vif(exog, var_names=None):
    """Calcula Variance Inflation Factor para cada predictor.

    VIF > 5 → multicolinealidad moderada.
    VIF > 10 → multicolinealidad severa.
    """
    from statsmodels.regression.linear_model import OLS
    from statsmodels.tools import add_constant

    X = add_constant(exog)
    n_vars = X.shape[1]
    vif_values = []

    for i in range(n_vars):
        y = X[:, i]
        X_others = np.column_stack([X[:, j] for j in range(n_vars) if j != i])
        try:
            model = OLS(y, X_others).fit()
            r2 = model.rsquared
            vif = 1.0 / (1.0 - r2) if r2 < 1 else np.inf
        except Exception:
            vif = np.inf
        vif_values.append(vif)

    names = ["const"] + (var_names if var_names else [f"V{i+1}" for i in range(n_vars - 1)])
    return dict(zip(names, vif_values))


def diagnostic_cophenetic(Z, dist_matrix):
    """Correlación cofenética: mide qué tan bien el dendrograma preserva las distancias.

    Retorna el coeficiente de correlación de Pearson entre distancias originales
    y distancias cofenéticas. > 0.7 → buena representación.
    """
    from scipy.cluster.hierarchy import cophenet
    cof, _ = cophenet(Z, dist_matrix)
    return cof


def interpretar_kmo(kmo):
    if kmo >= 0.9:
        return "Excelente"
    elif kmo >= 0.8:
        return "Meritorio"
    elif kmo >= 0.7:
        return "Intermedio"
    elif kmo >= 0.6:
        return "Mediocre"
    elif kmo >= 0.5:
        return "Pobre"
    else:
        return "Inaceptable"


def interpretar_dw(dw):
    if 1.5 < dw < 2.5:
        return "Sin autocorrelación relevante"
    elif dw < 1.0 or dw > 3.0:
        return "Posible autocorrelación severa"
    elif dw < 1.5:
        return "Posible autocorrelación positiva"
    else:
        return "Posible autocorrelación negativa"


METODOS_EXTRACCION = ["MACERACIÓN", "SOXHLET", "ULTRASONIDO"]
METODO_A_NORMALIZADO = {
    "MACERACIÓN": "maceracion",
    "SOXHLET": "soxhlet",
    "ULTRASONIDO": "ultrasonido",
}

# Mapeo de concentraciones: 
#   clave = código en la hoja (C1, C2, C3, C4)
#   valor = concentración en mg/mL
CONCENTRACIONES = {"C1": 5.0, "C2": 1.0, "C3": 0.2, "C4": 0.0}

# ─── Configuración de columnas por hoja ─────────────────────────────
# Cada hoja tiene una estructura distinta. Este diccionario describe
# dónde encontrar cada pieza de información en cada hoja (0-indexado).
#
# Para RADIO (crecimiento micelial):
#   - columna_aislado: columna con el código del aislado
#   - columna_grupo: columna con el grupo experimental
#   - columna_replica: columna con "RÉPLICA 1/2/3"
#   - columnas_conc: dict {concentración_mg_ml: índice_columna}
#   - columna_inh: columna con %INH precalculado (o None)
#   - columna_extra: columna extra sin nombre (o None)
#
# Para CONIDIAS (producción de conidias):
#   misma estructura, pero apuntando al bloque de conidias.

HOJAS_CONFIG = {
    "MACERACIÓN": {
        "columna_aislado": 0,
        "columna_grupo": 1,
        "radio": {
            "columna_replica": 2,
            "columnas_conc": {5.0: 3, 1.0: 4, 0.2: 5, 0.0: 6},
            "columna_inh": 7,
            "columna_extra": 8,
        },
        "conidias": {
            "columna_replica": 9,
            "columnas_conc": {5.0: 10, 1.0: 11, 0.2: 12, 0.0: 13},
            "columna_inh": 14,
            "columna_extra": 15,
        },
    },
    "SOXHLET": {
        "columna_aislado": 0,
        "columna_grupo": 1,
        "radio": {
            "columna_replica": 2,
            "columnas_conc": {5.0: 3, 0.0: 4},
            "columna_inh": 5,
            "columna_extra": 6,
        },
        "conidias": {
            "columna_replica": 7,
            "columnas_conc": {5.0: 8, 0.0: 9},
            "columna_inh": 10,
            "columna_extra": 11,
        },
    },
    "ULTRASONIDO": {
        # ATENCIÓN: ULTRASONIDO tiene una columna '#' extra al inicio
        # que desplaza todo. Aislado está en col 1, grupo en col 2.
        "columna_aislado": 1,
        "columna_grupo": 2,
        "radio": {
            "columna_replica": 3,
            "columnas_conc": {5.0: 4, 0.0: 5},
            "columna_inh": 6,
            "columna_extra": 7,
        },
        "conidias": {
            "columna_replica": 8,
            "columnas_conc": {5.0: 9, 0.0: 10},
            "columna_inh": 11,
            "columna_extra": 12,
        },
    },
}

# ─── Mapeo de normalización de aislados ─────────────────────────────
NORMALIZAR_AISLADOS = {
    "FU2 (UCMU21)": "FU2 (UCMU21)",
    "FU2": "FU2 (UCMU21)",                # unificar con la versión completa
    "FUSARIUM JULIAN H20": "FUSARIUM JULIAN H20",
    "FUSARIUM MARCE 1.2": "FUSARIUM MARCE 1.2",
}

# ─── Funciones compartidas ──────────────────────────────────────────


def leer_hoja_sin_encabezado(hoja):
    """Lee una hoja del Excel sin tratar ninguna fila como encabezado."""
    return pd.read_excel(
        EXCEL_ORIGINAL, sheet_name=hoja, engine="openpyxl", header=None
    )


def extraer_etiquetas_conc(hoja):
    """Lee la fila 2 de una hoja método y extrae las etiquetas de concentración.

    Retorna un dict {índice_columna: etiqueta_str} solo para las columnas
    que contienen etiquetas como 'C1 \\n(5 mg/ml)', etc.
    """
    df = leer_hoja_sin_encabezado(hoja)
    etiquetas = {}
    for col in range(df.shape[1]):
        val = df.iloc[2, col]
        if pd.notna(val) and isinstance(val, str) and ("mg/ml" in val or "CTROL" in val):
            etiquetas[col] = val.strip()
    return etiquetas


def parsear_bloque_metodo(hoja, bloque="radio"):
    """Parsea el bloque RADIO o CONIDIAS de una hoja método.

    Parámetros
    ----------
    hoja : str
        Nombre de la hoja (MACERACIÓN, SOXHLET, ULTRASONIDO).
    bloque : str
        'radio' o 'conidias'.

    Retorna
    -------
    pd.DataFrame con columnas:
        aislado_id, metodo_extraccion, grupo_experimental, replica_biologica,
        concentracion_mg_ml, crecimiento_mm (o conidias_log10),
        porcentaje_inhibicion_hoja, extra_medicion, hoja_origen, fila_origen
    """
    cfg = HOJAS_CONFIG[hoja]
    bloque_cfg = cfg[bloque]
    df = leer_hoja_sin_encabezado(hoja)

    metodo = METODO_A_NORMALIZADO[hoja]
    col_aislado = cfg["columna_aislado"]
    col_grupo = cfg["columna_grupo"]
    col_replica = bloque_cfg["columna_replica"]
    col_extra = bloque_cfg.get("columna_extra")

    # ── 1. Extraer solo las filas de datos (desde fila 3 en adelante) ──
    datos = df.iloc[3:].copy()
    filas_inicio = datos.index[0]  # índice real de la primera fila de datos

    # ── 2. Forward-fill: aislado y grupo solo aparecen en RÉPLICA 1 ──
    datos[col_aislado] = datos[col_aislado].replace("#", np.nan).ffill()
    datos[col_grupo] = datos[col_grupo].ffill()

    # ── 3. Extraer réplica ──
    replica = datos[col_replica].str.extract(r"RÉPLICA (\d+)", expand=False)
    replica = pd.to_numeric(replica, errors="coerce")

    # ── 4. Pivotar concentraciones: ancho → largo ──
    filas = []
    for conc, col_idx in bloque_cfg["columnas_conc"].items():
        valores = pd.to_numeric(datos[col_idx], errors="coerce")

        for i in range(len(datos)):
            aislado = datos.iloc[i, col_aislado]
            if pd.isna(aislado):
                continue  # fila sin aislado (no debería pasar tras ffill)
            grupo = datos.iloc[i, col_grupo]
            rep = replica.iloc[i]
            valor = valores.iloc[i]

            # Extraer %INH precalculado de la hoja
            inh_col = bloque_cfg.get("columna_inh")
            inh_val = None
            if inh_col is not None:
                inh_val = pd.to_numeric(datos.iloc[i, inh_col], errors="coerce")

            # Extraer columna extra
            extra_val = None
            if col_extra is not None:
                extra_val = pd.to_numeric(datos.iloc[i, col_extra], errors="coerce")

            filas.append(
                {
                    "aislado_id": str(aislado).strip(),
                    "metodo_extraccion": metodo,
                    "grupo_experimental": str(grupo).strip() if pd.notna(grupo) else None,
                    "replica_biologica": int(rep) if pd.notna(rep) else None,
                    "concentracion_mg_ml": conc,
                    "crecimiento_mm" if bloque == "radio" else "conidias_log10": valor,
                    "porcentaje_inhibicion_hoja": inh_val if pd.notna(inh_val) else None,
                    "extra_medicion": extra_val if pd.notna(extra_val) else None,
                    "es_control": conc == 0.0,
                    "hoja_origen": hoja,
                    "fila_origen": filas_inicio + i,
                }
            )

    resultado = pd.DataFrame(filas)
    return resultado


def parsear_crecimiento(hoja):
    """Parsea el bloque RADIO de una hoja método → crecimiento micelial."""
    return parsear_bloque_metodo(hoja, bloque="radio")


def parsear_conidias(hoja):
    """Parsea el bloque CONIDIAS de una hoja método → conidias."""
    df = parsear_bloque_metodo(hoja, bloque="conidias")
    # Renombrar columna específica de conidias
    df = df.rename(columns={"conidias_log10": "conidias_log10"})
    return df


def validar_rangos(df, columna, min_val, max_val, nombre_col):
    """Valida que una columna numérica esté dentro de un rango esperado.

    Retorna un DataFrame con los valores fuera de rango.
    """
    fuera = df[(df[columna] < min_val) | (df[columna] > max_val)]
    if len(fuera) > 0:
        print(
            f"  ⚠ {len(fuera)} valores fuera de rango [{min_val}, {max_val}] "
            f"en '{nombre_col}'"
        )
    return fuera
