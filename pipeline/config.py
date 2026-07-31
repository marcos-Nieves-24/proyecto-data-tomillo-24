"""Configuracion central de la pipeline reproducible.

Define rutas de entrada/salida (incluido el Excel crudo del laboratorio),
constantes cientificas, semillas aleatorias y utilidades de exportacion
(figuras PNG/PDF, tablas CSV y libros Excel) compartidas por todos los
modulos.

Los directorios de salida se crean automaticamente al importar el modulo.
"""

from __future__ import annotations

import unicodedata
from pathlib import Path

import matplotlib
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Rutas
# ---------------------------------------------------------------------------

RAIZ = Path(__file__).resolve().parent.parent

# Excel crudo del laboratorio: fuente de los controles C4 por aislado (hojas
# MACERACIÓN, SOXHLET y ULTRASONIDO). No debe modificarse.
EXCEL_CRUDO = RAIZ / "datos-proyectos tomillo-fusarium.xlsx"

EXCEL_TIDY = RAIZ / "resultados" / "database" / "consolidado_tidy.xlsx"
CSV_RENDIMIENTO = RAIZ / "resultados" / "database" / "rendimiento_extraccion.csv"

DIR_DATABASE = RAIZ / "resultados" / "database"
DIR_TABLAS = RAIZ / "resultados" / "tablas"
DIR_FIGURAS = RAIZ / "resultados" / "figuras"
DIR_REPORTES = RAIZ / "resultados" / "reportes"
DIR_EXCEL = RAIZ / "resultados" / "excel"

MASTER_CSV = DIR_DATABASE / "master_dataset_tomillo_fusarium.csv"
MASTER_XLSX = DIR_DATABASE / "master_dataset_tomillo_fusarium.xlsx"
DICCIONARIO_MD = DIR_TABLAS / "diccionario_datos.md"
INFORME_MD = DIR_REPORTES / "informe_final.md"
INFORME_HTML = DIR_REPORTES / "informe_final.html"
EXCEL_RESUMEN = DIR_EXCEL / "resumen_analisis.xlsx"

for _directorio in (DIR_DATABASE, DIR_TABLAS, DIR_FIGURAS, DIR_REPORTES, DIR_EXCEL):
    _directorio.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Constantes cientificas y tecnicas
# ---------------------------------------------------------------------------

SEMILLA_ALEATORIA = 42
DPI = 300
CONCENTRACION_UNICA_MG_ML = 5.0  # unica concentracion ensayada (5 mg/mL)

METODOS = ["maceracion", "soxhlet", "ultrasonido"]
METODO_LABEL = {
    "maceracion": "Maceración",
    "soxhlet": "Soxhlet",
    "ultrasonido": "Ultrasonido",
}
METODO_LABEL_INV = {v: k for k, v in METODO_LABEL.items()}

# Paleta daltónico-safe (Okabe-Ito): azul, naranja, verde.
PALETA_METODOS = {
    "maceracion": "#0072B2",
    "soxhlet": "#D55E00",
    "ultrasonido": "#009E73",
}
PALETA_FIGURAS = [PALETA_METODOS[m] for m in METODOS]

TAMANO_FIG = (9, 5.5)

COLUMNAS_BIOENSAYO = [
    "metodo_extraccion",
    "aislamiento",
    "replica",
    "crecimiento_micelial_mm",
    "porcentaje_inhibicion_micelial",
    "conidias_log10_ml",
    "porcentaje_inhibicion_conidias",
]

VARIABLES_RESPUESTA = COLUMNAS_BIOENSAYO[3:]

VARIABLE_LABEL = {
    "crecimiento_micelial_mm": "Crecimiento micelial (mm)",
    "porcentaje_inhibicion_micelial": "Inhibición micelial (%)",
    "conidias_log10_ml": "Conidias (log10/mL)",
    "porcentaje_inhibicion_conidias": "Inhibición de conidias (%)",
}

VARIABLE_UNIDAD = {
    "crecimiento_micelial_mm": "mm",
    "porcentaje_inhibicion_micelial": "%",
    "conidias_log10_ml": "log10(conidias/mL)",
    "porcentaje_inhibicion_conidias": "%",
}

plt.rcParams.update({
    "figure.dpi": DPI,
    "savefig.dpi": DPI,
    "font.size": 11,
    "axes.titlesize": 12,
    "axes.labelsize": 11,
    "legend.fontsize": 10,
    "axes.grid": True,
    "grid.alpha": 0.35,
})


# ---------------------------------------------------------------------------
# Utilidades de exportacion
# ---------------------------------------------------------------------------


def save_figure_pub(fig, nombre, carpeta=None, titulo=None, dpi=DPI):
    """Guarda una figura como PNG y PDF (300 dpi) aptos para publicacion.

    Si ``titulo`` se provee, guarda primero la version con titulo y luego una
    variante ``_clean`` (sin titulo) para insercion directa en articulos.
    """
    carpeta = Path(carpeta) if carpeta is not None else DIR_FIGURAS
    carpeta.mkdir(parents=True, exist_ok=True)
    base = carpeta / nombre
    if titulo:
        fig.suptitle(titulo)
    fig.tight_layout(rect=(0, 0, 1, 0.95))
    fig.savefig(base.with_suffix(".png"), dpi=dpi, bbox_inches="tight")
    fig.savefig(base.with_suffix(".pdf"), dpi=dpi, bbox_inches="tight")
    if titulo:
        fig.suptitle("")
        limpio = carpeta / f"{nombre}_clean"
        fig.savefig(limpio.with_suffix(".png"), dpi=dpi, bbox_inches="tight")
        fig.savefig(limpio.with_suffix(".pdf"), dpi=dpi, bbox_inches="tight")
    plt.close(fig)
    return base.with_suffix(".png")


def guardar_tabla(df, nombre, index=True):
    """Guarda un DataFrame como CSV en ``resultados/tablas``."""
    ruta = DIR_TABLAS / f"{nombre}.csv"
    df.to_csv(ruta, index=index, encoding="utf-8")
    return ruta


def exportar_excel(hojas, ruta=EXCEL_RESUMEN):
    """Exporta un dict de DataFrames a un libro .xlsx con formato simple."""
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        for hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=str(hoja)[:31], index=False)
    _formatear_excel(ruta)
    return ruta


def _formatear_excel(ruta):
    """Aplica formato simple: encabezados en negrita, congelar fila y 3 decimales."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(ruta)
    encabezado = PatternFill("solid", fgColor="DDEBF7")
    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.font = Font(bold=True, color="1F3864")
            cell.fill = encabezado
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        for fila in ws.iter_rows(min_row=2):
            for cell in fila:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000"
    wb.save(ruta)


def fijar_semilla(seed=SEMILLA_ALEATORIA):
    """Fija las semillas aleatorias de numpy, random y matplotlib."""
    import random

    np.random.seed(seed)
    random.seed(seed)
    return seed


def normalizar_texto(s: str) -> str:
    """Normaliza un texto: strip, minusculas y sin acentos."""
    return _sin_acentos(s.strip().lower())


def _sin_acentos(s: str) -> str:
    """Elimina acentos de un texto mediante descomposicion Unicode."""
    if not isinstance(s, str):
        return s
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch)
    )
